import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image

# Open an existing Excel file
wb = openpyxl.load_workbook("random_monthly_report.xlsx")

# Create a summary sheet for charts
summary_sheet = wb.create_sheet("Summary")

# Iterate through all sheets in the workbook
for sheet_name in wb.sheetnames:
    # Get the current sheet
    sheet = wb[sheet_name]

    # Get the maximum number of columns in the sheet
    max_col = sheet.max_column

    # Iterate through all columns (starting from the second column)
    charts = ['Total ore carried', 'Total distance (km)', 'Working days', 'Number of trips', 'Average cycle time']
    for col in range(1, max_col + 1):
        if sheet.cell(row=1, column=col).value in charts:
            # Create data for plotting (using the first column and the current column)
            data = Reference(sheet, min_col=col, min_row=2, max_col=col, max_row=sheet.max_row)

            # Create labels for the x-axis (using the first column)
            labels = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)

            # Create an object of BarChart class
            chart = BarChart()

            # Adding data to the Bar chart object
            chart.add_data(data, titles_from_data=True)

            # Add labels to the x-axis
            chart.set_categories(labels)

            # Set the title of the chart
            chart.title = f"BAR-CHART - {sheet.cell(row=1,column=col).value} - Driver Id"

            # Set the title of the x-axis (column 1)
            chart.x_axis.title = sheet.cell(row=1, column=1).value

            # Set the title of the y-axis (current column)
            chart.y_axis.title = sheet.cell(row=1, column=col).value

            # Add chart to the summary sheet (anchored to a new position for each chart)
            anchor_cell = f"A{1 + (col - 2) * 20}"
            summary_sheet.add_chart(chart, anchor_cell)

# Save the modified workbook
wb.save("barCharts_with_summary.xlsx")